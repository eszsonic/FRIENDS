#!/usr/bin/env python3
"""Plot device-level mean puff-duration error from FRIENDS lifecycle data.

The machine reference duration is fixed (3.000 s in the supplied workbook), so
a conventional Bland-Altman difference-versus-mean plot is mathematically
coupled. This script instead produces a device-level forest plot. By default,
the constant reference duration is inferred and verified from the workbook.

For device v, the plotted error is

    reference duration - sum(recorded puff time) / sum(counted puffs)

where sums are taken over all puffing sessions for that device. Thus, the
duration analysis is conditional on FRIENDS-counted events; puff detection
performance should be reported separately using precision and recall.

Device whiskers are session-clustered 95% confidence intervals for the ratio
of total recorded time to total counted puffs. The horizontal dashed limits are
descriptive mean +/- 1.96 SD limits across the observed device means; they are
not limits of agreement for individual puffs. Devices with only one session 
are excluded as their confidence intervals cannot be estimated.

Requires: openpyxl, numpy, scipy, and matplotlib.

Example:
    python plot_duration_error_by_device.py \
        Lifecycle-testing-FRIENDS-Data-V5.xlsx \
        --output mean_puff_duration_error_by_device.png
"""

from __future__ import annotations

import argparse
import math
from collections import OrderedDict
from dataclasses import dataclass
from numbers import Real
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from scipy import stats


DEFAULT_WORKBOOK = "Lifecycle-testing-FRIENDS-Data-V5.xlsx"
DEFAULT_EXCLUSIONS = ("Nord 2", "Logic Pro")


@dataclass(frozen=True)
class Session:
    counted_puffs: float
    recorded_time_s: float
    machine_duration_s: float
    worksheet_row: int


@dataclass(frozen=True)
class DeviceEstimate:
    name: str
    sessions: tuple[Session, ...]
    mean_error_s: float
    ci_low_s: float
    ci_high_s: float

    @property
    def n_sessions(self) -> int:
        return len(self.sessions)

    @property
    def counted_puffs(self) -> float:
        return sum(s.counted_puffs for s in self.sessions)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create a corrected device-level mean puff-duration error plot "
            "from the 'All Data' worksheet. Devices with a single session are dropped."
        )
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=Path(__file__).resolve().parent / DEFAULT_WORKBOOK,
        help="Input .xlsx workbook (default: workbook beside this script).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output figure path; extension controls format (default: PNG beside workbook).",
    )
    parser.add_argument(
        "--reference-duration",
        type=float,
        default=None,
        help=(
            "Machine reference duration in seconds. By default, infer it from "
            "'Actual Puff Time (s)' and require a single constant value."
        ),
    )
    exclusion_group = parser.add_mutually_exclusive_group()
    exclusion_group.add_argument(
        "--exclude",
        nargs="+",
        default=None,
        metavar="DEVICE",
        help="Device names to exclude (default: Nord 2 and Logic Pro).",
    )
    exclusion_group.add_argument(
        "--include-all-devices",
        action="store_true",
        help="Include all devices, overriding the default exclusions.",
    )
    parser.add_argument(
        "--bootstrap-mode",
        choices=("sessions", "hierarchical"),
        default="sessions",
        help=(
            "CI for the overall mean bias: resample sessions within fixed devices "
            "(default), or resample devices and sessions (hierarchical)."
        ),
    )
    parser.add_argument(
        "--bootstrap-replicates",
        type=int,
        default=10_000,
        help="Number of bootstrap replicates for the overall-bias CI (default: 10000).",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=12_345,
        help="Random seed for reproducibility (default: 12345).",
    )
    parser.add_argument(
        "--order",
        choices=("error", "workbook"),
        default="error",
        help="Order devices by mean error or workbook order (default: error).",
    )
    parser.add_argument("--dpi", type=int, default=300, help="Output DPI (default: 300).")
    parser.add_argument("--show", action="store_true", help="Display the plot interactively.")
    return parser.parse_args()


def _find_header(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, dict[str, int]]:
    required = {
        "Vape",
        "Puff Session",
        "Actual Puff Time (s)",
        "Puffs Counted",
        "Recorded Puff Time (s)",
    }
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True), start=1
    ):
        columns = {str(value).strip(): i + 1 for i, value in enumerate(row) if value is not None}
        if required.issubset(columns):
            return row_number, columns
    raise ValueError(f"Could not find required headers in worksheet {ws.title!r}: {sorted(required)}")


def load_sessions(workbook: Path) -> OrderedDict[str, tuple[Session, ...]]:
    """Read session rows dynamically from the workbook's 'All Data' worksheet."""
    if not workbook.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook}")

    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    if "All Data" not in wb.sheetnames:
        raise KeyError("Workbook does not contain an 'All Data' worksheet.")

    ws = wb["All Data"]
    header_row, columns = _find_header(ws)
    grouped: OrderedDict[str, list[Session]] = OrderedDict()
    current_device: str | None = None

    for row_number, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        device_value = row[columns["Vape"] - 1]
        session_value = row[columns["Puff Session"] - 1]

        if device_value is not None and str(device_value).strip():
            current_device = str(device_value).strip()

        # Numeric values identify actual session rows; rows labeled "Overall" are skipped.
        if not isinstance(session_value, Real) or isinstance(session_value, bool):
            continue
        if current_device is None:
            raise ValueError(f"Session row {row_number} has no device name to forward-fill.")

        counted = row[columns["Puffs Counted"] - 1]
        recorded = row[columns["Recorded Puff Time (s)"] - 1]
        machine_duration = row[columns["Actual Puff Time (s)"] - 1]
        if not all(
            isinstance(value, Real)
            for value in (counted, recorded, machine_duration)
        ):
            raise ValueError(
                f"Row {row_number} ({current_device}) lacks numeric cached values for "
                "Actual Puff Time, Puffs Counted, or Recorded Puff Time. "
                "Recalculate and save the workbook."
            )
        if counted <= 0:
            raise ValueError(
                f"Row {row_number} ({current_device}) has {counted} counted puffs; "
                "mean recorded duration is undefined."
            )

        grouped.setdefault(current_device, []).append(
            Session(float(counted), float(recorded), float(machine_duration), row_number)
        )

    wb.close()
    if not grouped:
        raise ValueError("No numeric puffing-session rows were found in 'All Data'.")
    return OrderedDict((name, tuple(values)) for name, values in grouped.items())


def estimate_device(
    name: str, sessions: tuple[Session, ...], reference_duration_s: float
) -> DeviceEstimate:
    """Estimate a device's pooled mean error and session-cluster ratio CI."""
    x = np.asarray([s.counted_puffs for s in sessions], dtype=float)
    y = np.asarray([s.recorded_time_s for s in sessions], dtype=float)
    ratio = y.sum() / x.sum()
    error = reference_duration_s - ratio

    if len(sessions) < 2:
        ci_low = ci_high = math.nan
    else:
        residual = y - ratio * x
        variance = (
            len(sessions)
            * np.square(residual).sum()
            / ((len(sessions) - 1) * np.square(x.sum()))
        )
        half_width = stats.t.ppf(0.975, df=len(sessions) - 1) * math.sqrt(variance)
        ci_low, ci_high = error - half_width, error + half_width

    return DeviceEstimate(name, sessions, error, ci_low, ci_high)


def _resampled_device_error(
    sessions: tuple[Session, ...], reference_duration_s: float, rng: np.random.Generator
) -> float:
    indexes = rng.integers(0, len(sessions), len(sessions))
    counted = sum(sessions[i].counted_puffs for i in indexes)
    recorded = sum(sessions[i].recorded_time_s for i in indexes)
    return reference_duration_s - recorded / counted


def bootstrap_bias_ci(
    estimates: list[DeviceEstimate],
    reference_duration_s: float,
    mode: str,
    replicates: int,
    seed: int,
) -> tuple[float, float]:
    """Percentile CI for the equal-device-weighted overall mean bias."""
    if replicates < 1_000:
        raise ValueError("Use at least 1000 bootstrap replicates.")

    rng = np.random.default_rng(seed)
    n_devices = len(estimates)
    bootstrap = np.empty(replicates, dtype=float)

    for b in range(replicates):
        if mode == "sessions":
            # Fixed observed devices; only sessions within each device are resampled.
            errors = [
                _resampled_device_error(e.sessions, reference_duration_s, rng)
                for e in estimates
            ]
        elif mode == "hierarchical":
            # Generalize to new devices and new sessions.
            device_indexes = rng.integers(0, n_devices, n_devices)
            errors = [
                _resampled_device_error(
                    estimates[i].sessions, reference_duration_s, rng
                )
                for i in device_indexes
            ]
        else:  # Protected by argparse, retained for programmatic use.
            raise ValueError(f"Unknown bootstrap mode: {mode}")
        bootstrap[b] = float(np.mean(errors))

    low, high = np.percentile(bootstrap, [2.5, 97.5])
    return float(low), float(high)


def make_plot(
    estimates: list[DeviceEstimate],
    reference_duration_s: float,
    bias_ci: tuple[float, float],
    bootstrap_mode: str,
    output: Path,
    dpi: int,
    order: str,
) -> tuple[float, float, float]:
    """Create and save the corrected device-level forest plot."""
    if order == "error":
        plotted = sorted(estimates, key=lambda item: item.mean_error_s, reverse=True)
    else:
        plotted = list(estimates)

    errors = np.asarray([e.mean_error_s for e in estimates], dtype=float)
    overall_bias = float(errors.mean())
    device_sd = float(errors.std(ddof=1))
    limit_low = overall_bias - 1.96 * device_sd
    limit_high = overall_bias + 1.96 * device_sd

    n_devices = len(plotted)
    # Scale height dynamically, but fix width to 3.5 inches (IEEE single column)
    figure_height = max(3.5, 0.18 * n_devices + 1.5)
    fig, ax = plt.subplots(figsize=(3.5, figure_height))

    # Draw summary references behind the device estimates.
    ax.axvspan(bias_ci[0], bias_ci[1], color="#6B7280", alpha=0.16, zorder=0)
    ax.axvline(0.0, color="#6B7280", linestyle=":", linewidth=1.0, zorder=1)
    ax.axvline(overall_bias, color="#111827", linewidth=1.2, zorder=1)
    ax.axvline(limit_low, color="#DC2626", linestyle="--", linewidth=1.0, zorder=1)
    ax.axvline(limit_high, color="#DC2626", linestyle="--", linewidth=1.0, zorder=1)

    y_positions = np.arange(n_devices)
    for y_position, estimate in zip(y_positions, plotted):
        # All plotted devices now have >1 session, so CI is always finite
        ax.errorbar(
            estimate.mean_error_s,
            y_position,
            xerr=np.asarray(
                [
                    [estimate.mean_error_s - estimate.ci_low_s],
                    [estimate.ci_high_s - estimate.mean_error_s],
                ]
            ),
            fmt="o",
            markersize=3.5,
            color="#1F77B4",
            ecolor="#4B5563",
            elinewidth=0.8,
            capsize=1.5,
            capthick=0.8,
            zorder=3,
        )

    ax.set_yticks(y_positions)
    ax.set_yticklabels([e.name for e in plotted], fontsize=7)
    ax.invert_yaxis()
    ax.set_xlabel("Mean duration error: reference − FRIENDS (s)", fontsize=8)
    ax.set_title(
        "Mean puff-duration error by ENDS device\n"
        f"Reference duration = {reference_duration_s:.3f} s",
        fontsize=9,
        pad=8,
    )
    ax.grid(axis="x", color="#D1D5DB", linewidth=0.6, linestyle=":")
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    finite_ci_values = [
        value for e in estimates for value in (e.ci_low_s, e.ci_high_s) if math.isfinite(value)
    ]
    x_values = [*errors.tolist(), *finite_ci_values, limit_low, limit_high, 0.0]
    x_min, x_max = min(x_values), max(x_values)
    margin = max(0.02, 0.05 * (x_max - x_min))
    ax.set_xlim(x_min - margin, x_max + margin)

    bootstrap_label = (
        f"Bias 95% CI: {bias_ci[0]:+.2f} to {bias_ci[1]:+.2f} s (session boot)"
        if bootstrap_mode == "sessions"
        else f"Bias 95% CI: {bias_ci[0]:+.2f} to {bias_ci[1]:+.2f} s (hierarchical boot)"
    )
    legend_handles = [
        Line2D([0], [0], marker="o", color="#4B5563", markerfacecolor="#1F77B4", markersize=3.5, linewidth=0.8, label="Device mean with 95% CI"),
        Line2D([0], [0], color="#111827", linewidth=1.2, label=f"Mean bias = {overall_bias:+.2f} s"),
        Patch(facecolor="#6B7280", alpha=0.16, label=bootstrap_label),
        Line2D([0], [0], color="#DC2626", linestyle="--", linewidth=1.0, label=f"Device limits: {limit_low:+.2f} to {limit_high:+.2f} s"),
    ]
    
    # Place legend in the upper-left corner inside the plot axes
    ax.legend(
        handles=legend_handles, 
        loc="upper left", 
        fontsize=5, 
        frameon=True,         # Turned back on so the background isn't transparent over the lines
        framealpha=0.95,      # Slight transparency
        edgecolor="#D1D5DB"
    )

    # Note: For IEEE formatting, consider moving this explanatory text to the manuscript's figure caption.
    # Adjusted 'xy' to tuck it closer to the x-axis now that the legend is moved.
    ax.annotate(
        "Each point = reference − (Σ FRIENDS time / Σ FRIENDS puffs).\n"
        "Whiskers: session-clustered CIs. Dashed lines: descriptive limits.",
        xy=(0.5, -0.2), xycoords="axes fraction", ha="center", va="top",
        fontsize=6.5, color="#4B5563",
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output, dpi=dpi, bbox_inches="tight", facecolor="white")
    return overall_bias, limit_low, limit_high


def main() -> None:
    args = parse_args()
    if args.dpi < 72:
        raise ValueError("DPI must be at least 72.")

    workbook = args.workbook.resolve()
    output = (
        args.output.resolve()
        if args.output is not None
        else workbook.with_name(f"{workbook.stem}_duration_error_by_device.png")
    )
    if args.include_all_devices:
        exclusions: set[str] = set()
    elif args.exclude is not None:
        exclusions = {name.casefold() for name in args.exclude}
    else:
        exclusions = {name.casefold() for name in DEFAULT_EXCLUSIONS}

    all_sessions = load_sessions(workbook)
    included_sessions = OrderedDict(
        (name, sessions)
        for name, sessions in all_sessions.items()
        if name.casefold() not in exclusions
    )
    excluded_found = [
        name for name in all_sessions if name.casefold() in exclusions
    ]
    if not included_sessions:
        raise ValueError("All devices were excluded; there is nothing to plot.")

    machine_durations = np.asarray(
        [
            session.machine_duration_s
            for sessions in included_sessions.values()
            for session in sessions
        ],
        dtype=float,
    )
    if args.reference_duration is None:
        reference_duration = float(machine_durations[0])
        if not np.allclose(machine_durations, reference_duration, rtol=0.0, atol=1e-9):
            unique_values = np.unique(machine_durations)
            raise ValueError(
                "The workbook does not contain one constant Actual Puff Time. "
                f"Observed values: {unique_values.tolist()}. A constant-reference "
                "device-error plot is not appropriate without stratification."
            )
    else:
        reference_duration = float(args.reference_duration)
        if not np.allclose(machine_durations, reference_duration, rtol=0.0, atol=1e-9):
            raise ValueError(
                f"Requested reference duration {reference_duration:g} s does not match "
                "all included 'Actual Puff Time (s)' values in the workbook."
            )
    if reference_duration <= 0:
        raise ValueError("Reference duration must be positive.")

    # Filter out devices with only one session
    estimates = [
        estimate_device(name, sessions, reference_duration)
        for name, sessions in included_sessions.items()
        if len(sessions) > 1
    ]
    
    if not estimates:
        raise ValueError("No devices with >1 session remain to plot.")

    bias_ci = bootstrap_bias_ci(
        estimates,
        reference_duration,
        args.bootstrap_mode,
        args.bootstrap_replicates,
        args.seed,
    )
    overall_bias, limit_low, limit_high = make_plot(
        estimates,
        reference_duration,
        bias_ci,
        args.bootstrap_mode,
        output,
        args.dpi,
        args.order,
    )

    n_sessions = sum(e.n_sessions for e in estimates)
    n_counted = sum(e.counted_puffs for e in estimates)
    print(f"Workbook: {workbook}")
    print(f"Output: {output}")
    print(
        f"Included: {len(estimates)} devices, {n_sessions} sessions, "
        f"{n_counted:.0f} FRIENDS-counted puffs"
    )
    print(f"Excluded: {', '.join(excluded_found) if excluded_found else 'none'}")
    print(f"Equal-device mean bias: {overall_bias:+.6f} s")
    print(
        f"Overall-bias 95% CI ({args.bootstrap_mode} bootstrap): "
        f"[{bias_ci[0]:+.6f}, {bias_ci[1]:+.6f}] s"
    )
    print(
        "Descriptive limits across device means (mean +/- 1.96 SD): "
        f"[{limit_low:+.6f}, {limit_high:+.6f}] s"
    )

    if args.show:
        plt.show()
    else:
        plt.close("all")


if __name__ == "__main__":
    main()