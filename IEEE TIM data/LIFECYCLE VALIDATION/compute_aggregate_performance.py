"""
Device-cluster bootstrap for FRIENDS lifecycle bench testing.

Each ENDS device is one bootstrap cluster. Metrics are recomputed from pooled
TP, FP, FN, puff-count error, and puff-time error totals in every replicate.

Usage
-----
python compute_aggregate_performance.py [workbook.xlsx]
"""

from __future__ import annotations

import argparse
import math
from collections import OrderedDict
from numbers import Real
from pathlib import Path

import numpy as np
import openpyxl
from scipy import stats


DEFAULT_XLSX = "Lifecycle-testing-FRIENDS-Data-V5.xlsx"
DEFAULT_BOOTSTRAP_REPLICATES = 10_000
DEFAULT_SEED = 12_345

REQUIRED_HEADERS = (
    "Vape",
    "Puff Session",
    "Number of Puffs",
    "Total Puff Time (s)",
    "Puffs Counted",
    "Recorded Puff Time (s)",
    "TP",
    "FP",
    "FN",
)


def _is_finite_number(value: object) -> bool:
    return (
        isinstance(value, Real)
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    )


def _resolve_workbook(requested_path: Path | None) -> Path:
    """Resolve an explicit path or the standard workbook beside the script/CWD."""
    if requested_path is not None:
        path = requested_path.expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path

    candidates = [
        Path(__file__).resolve().with_name(DEFAULT_XLSX),
        Path.cwd() / DEFAULT_XLSX,
    ]
    for path in candidates:
        if path.is_file():
            return path.resolve()

    # This also permits browser/download suffixes such as "(1)" when there is
    # exactly one unambiguous V5 workbook in the current directory.
    matches = sorted(Path.cwd().glob("Lifecycle-testing-FRIENDS-Data-V5*.xlsx"))
    if len(matches) == 1:
        return matches[0].resolve()

    detail = ""
    if len(matches) > 1:
        detail = " Multiple matching workbooks were found; specify one explicitly."
    raise FileNotFoundError(
        f"Could not find {DEFAULT_XLSX}. Pass the workbook path on the command line."
        + detail
    )


def _header_map(worksheet, header_row: int = 2) -> dict[str, int]:
    """Return 1-based column indices and reject a missing/ambiguous schema."""
    locations: dict[str, list[int]] = {name: [] for name in REQUIRED_HEADERS}
    for cell in worksheet[header_row]:
        if cell.value in locations:
            locations[cell.value].append(cell.column)

    missing = [name for name, columns in locations.items() if not columns]
    duplicates = [name for name, columns in locations.items() if len(columns) > 1]
    if missing or duplicates:
        parts = []
        if missing:
            parts.append("missing headers: " + ", ".join(missing))
        if duplicates:
            parts.append("duplicate headers: " + ", ".join(duplicates))
        raise ValueError("Invalid 'All Data' schema (" + "; ".join(parts) + ")")

    return {name: columns[0] for name, columns in locations.items()}


def load_devices(path: Path) -> tuple[np.ndarray, list[str]]:
    """
    Return one row per device:
    [TP, FP, FN, absolute count error, delivered puffs,
     absolute puff-time error, programmed puff time].

    Device blocks and session rows are read directly from ``All Data``. This
    avoids parsing formulas in ``Data Summary``, where single-session devices
    use a single-cell reference instead of a cell range.
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if "All Data" not in workbook.sheetnames:
            raise ValueError("Workbook does not contain an 'All Data' worksheet")

        worksheet = workbook["All Data"]
        columns = _header_map(worksheet)
        devices: OrderedDict[str, np.ndarray] = OrderedDict()
        session_counts: OrderedDict[str, int] = OrderedDict()
        current_device: str | None = None

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=3, values_only=True), start=3
        ):
            device_cell = row[columns["Vape"] - 1]
            if device_cell is not None and str(device_cell).strip():
                current_device = str(device_cell).strip()

            session = row[columns["Puff Session"] - 1]
            if not _is_finite_number(session):
                # Skips blank/separator rows and each device's "Overall" row.
                continue
            if current_device is None:
                raise ValueError(
                    f"Session row {row_number} has no preceding device name"
                )

            field_values = {
                name: row[columns[name] - 1]
                for name in REQUIRED_HEADERS
                if name not in ("Vape", "Puff Session")
            }
            invalid = [
                name for name, value in field_values.items()
                if not _is_finite_number(value)
            ]
            if invalid:
                raise ValueError(
                    f"Non-numeric or missing data in 'All Data' row {row_number}: "
                    + ", ".join(invalid)
                    + ". Recalculate and save the workbook if formula values are stale."
                )

            delivered = float(field_values["Number of Puffs"])
            programmed_time = float(field_values["Total Puff Time (s)"])
            counted = float(field_values["Puffs Counted"])
            recorded_time = float(field_values["Recorded Puff Time (s)"])
            tp = float(field_values["TP"])
            fp = float(field_values["FP"])
            fn = float(field_values["FN"])

            if min(delivered, programmed_time, counted, recorded_time, tp, fp, fn) < 0:
                raise ValueError(f"Negative measurement in 'All Data' row {row_number}")
            if delivered == 0 or programmed_time == 0:
                raise ValueError(
                    f"Zero denominator in 'All Data' row {row_number}"
                )
            if not math.isclose(tp + fn, delivered, rel_tol=0.0, abs_tol=1e-9):
                raise ValueError(
                    f"Inconsistent TP/FN counts in 'All Data' row {row_number}: "
                    f"TP + FN = {tp + fn:g}, delivered = {delivered:g}"
                )
            if not math.isclose(tp + fp, counted, rel_tol=0.0, abs_tol=1e-9):
                raise ValueError(
                    f"Inconsistent TP/FP counts in 'All Data' row {row_number}: "
                    f"TP + FP = {tp + fp:g}, counted = {counted:g}"
                )

            if current_device not in devices:
                devices[current_device] = np.zeros(7, dtype=float)
                session_counts[current_device] = 0

            devices[current_device] += np.array(
                [
                    tp,
                    fp,
                    fn,
                    abs(counted - delivered),
                    delivered,
                    abs(recorded_time - programmed_time),
                    programmed_time,
                ],
                dtype=float,
            )
            session_counts[current_device] += 1

        if not devices:
            raise ValueError("No numeric puff-session rows were found in 'All Data'")

        names = list(devices)
        no_sessions = [name for name in names if session_counts[name] == 0]
        if no_sessions:
            raise ValueError("Devices without sessions: " + ", ".join(no_sessions))
        return np.vstack(list(devices.values())), names
    finally:
        workbook.close()


def pooled(device_totals: np.ndarray) -> np.ndarray:
    """Compute pooled precision, recall, F1, count WMAPE, and time WMAPE."""
    tp, fp, fn, count_abs_error, puffs, time_abs_error, puff_time = (
        device_totals.sum(axis=0)
    )
    precision_denominator = tp + fp
    recall_denominator = tp + fn
    if min(precision_denominator, recall_denominator, puffs, puff_time) <= 0:
        raise ValueError("A pooled metric has a zero denominator")

    precision = tp / precision_denominator
    recall = tp / recall_denominator
    f1 = 0.0 if precision + recall == 0 else (
        2.0 * precision * recall / (precision + recall)
    )
    return np.array(
        [
            precision,
            recall,
            f1,
            count_abs_error / puffs,
            time_abs_error / puff_time,
        ]
    )


def bca_interval(
    theta_hat: float, bootstrap_values: np.ndarray, jackknife_values: np.ndarray
) -> tuple[float, float]:
    """Return a two-sided 95% bias-corrected and accelerated interval."""
    n_boot = len(bootstrap_values)
    proportion_less = np.mean(bootstrap_values < theta_hat)
    # Prevent infinite z0 when a finite bootstrap sample lies on one side only.
    proportion_less = np.clip(
        proportion_less, 0.5 / n_boot, 1.0 - 0.5 / n_boot
    )
    z0 = stats.norm.ppf(proportion_less)

    jackknife_mean = jackknife_values.mean()
    differences = jackknife_mean - jackknife_values
    acceleration_denominator = 6.0 * np.sum(differences**2) ** 1.5
    acceleration = (
        0.0
        if acceleration_denominator == 0
        else np.sum(differences**3) / acceleration_denominator
    )

    bounds = []
    for quantile in (0.025, 0.975):
        z = stats.norm.ppf(quantile)
        denominator = 1.0 - acceleration * (z0 + z)
        adjusted = stats.norm.cdf(z0 + (z0 + z) / denominator)
        adjusted = float(np.clip(adjusted, 0.0, 1.0))
        bounds.append(float(np.quantile(bootstrap_values, adjusted)))
    return bounds[0], bounds[1]


def run_analysis(path: Path, replicates: int, seed: int) -> None:
    if replicates < 1:
        raise ValueError("Bootstrap replicate count must be positive")

    device_totals, names = load_devices(path)
    n_devices = len(device_totals)
    if n_devices < 2:
        raise ValueError("At least two devices are required for cluster bootstrap")

    point_estimates = pooled(device_totals)
    rng = np.random.default_rng(seed)
    bootstrap = np.array(
        [
            pooled(device_totals[rng.integers(0, n_devices, n_devices)])
            for _ in range(replicates)
        ]
    )
    jackknife = np.array(
        [pooled(np.delete(device_totals, i, axis=0)) for i in range(n_devices)]
    )

    metrics = (
        "precision",
        "recall",
        "F1",
        "puff-count WMAPE",
        "puff-time WMAPE",
    )

    print(f"Workbook: {path}")
    print("Included devices: " + ", ".join(names))
    print(
        f"\nDevice-cluster bootstrap: n = {n_devices} devices, "
        f"B = {replicates:,}, seed = {seed}\n"
    )
    print("All metrics are fractions rounded to two decimal places.\n")
    print(f"{'metric':24}{'estimate':>10}{'percentile 95% CI':>26}{'BCa 95% CI':>26}")
    for i, label in enumerate(metrics):
        percentile_low, percentile_high = np.percentile(
            bootstrap[:, i], [2.5, 97.5]
        )
        bca_low, bca_high = bca_interval(
            point_estimates[i], bootstrap[:, i], jackknife[:, i]
        )
        estimate_text = format(point_estimates[i], ".2f")
        percentile_text = (
            f"[{format(percentile_low, '.2f')}, "
            f"{format(percentile_high, '.2f')}]"
        )
        bca_text = (
            f"[{format(bca_low, '.2f')}, "
            f"{format(bca_high, '.2f')}]"
        )
        print(
            f"{label:24}{estimate_text:>10}"
            f"{percentile_text:>26}{bca_text:>26}"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compute aggregate FRIENDS lifecycle metrics with a device-cluster bootstrap."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        help=f"Input workbook (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--bootstrap-replicates",
        type=int,
        default=DEFAULT_BOOTSTRAP_REPLICATES,
        help=f"Number of bootstrap replicates (default: {DEFAULT_BOOTSTRAP_REPLICATES:,})",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=DEFAULT_SEED,
        help=f"Random seed (default: {DEFAULT_SEED})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = _resolve_workbook(args.workbook)
    run_analysis(workbook_path, args.bootstrap_replicates, args.seed)


if __name__ == "__main__":
    main()
