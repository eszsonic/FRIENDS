#!/usr/bin/env python3
"""
Bland-Altman analysis of FRIENDS puff-time agreement on the lifecycle bench data,
styled to match the human laboratory figure (Video vs FRIENDS puff duration).

Two modes, because the bench reference is not the same kind of quantity as the
video-coded reference:

  --mode total    (default)  Total atomizer-activation time per device (s).
                             The reference quantity varies across devices
                             (555-3210 s), so a conventional Bland-Altman plot
                             with a proportional-bias test is valid.

  --mode perpuff             Mean puff duration (s), directly comparable in
                             units to the human figure's +0.52 s bias. The
                             puffing machine held every puff at exactly 3.000 s,
                             so the reference has zero variance. Difference and
                             mean are then perfectly coupled (mean = 3 - diff/2),
                             which forces r = -1 by construction. No
                             proportional-bias fit is drawn in this mode; the
                             bias and limits of agreement are shown against
                             device index instead.

Unit of analysis: the ENDS device (the cluster), matching the confidence-interval
convention used elsewhere in the paper. Use --level session for one point per
puffing session.

Usage:
    python bland_altman_bench.py FRIENDS-Data-V4.xlsx
    python bland_altman_bench.py FRIENDS-Data-V4.xlsx --mode perpuff
    python bland_altman_bench.py FRIENDS-Data-V4.xlsx --level session --all-devices
"""

import argparse
import sys

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from scipy import stats

# ----------------------------------------------------------------------------
# Style constants copied from the human laboratory figure
# ----------------------------------------------------------------------------
C_POINT = "#4a89b8"       # steel blue markers
C_POINT_EDGE = "#2f6c96"
C_BIAS = "black"          # solid bias line
C_LOA = "crimson"         # dashed limits of agreement
C_FIT = "orange"          # dash-dot proportional-bias fit
C_BIAS_BAND = "0.82"      # grey 95% CI band around the bias
C_LOA_BAND = "#f7dfe4"    # pink 95% CI bands around each LoA
Z = 1.96                  # multiplier defining the limits of agreement


def read_sessions(path):
    """Return (included_device_names, list of per-session dicts) from the workbook.

    'All Data' is laid out as blocks of puffing sessions per vape, with the vape
    name written only on the first row of the block and an 'Overall' summary row
    closing each block. Column C holds an integer session number on data rows,
    which is what distinguishes them from the 'Overall' rows.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ad, ds = wb["All Data"], wb["Data Summary"]

    # Devices retained for lifecycle analysis occupy the contiguous block above
    # the blank spacer row in 'Data Summary'; the four devices that lost
    # detectability sit below it.
    included, row = [], 3
    while ds.cell(row, 2).value is not None:
        included.append(str(ds.cell(row, 2).value).strip())
        row += 1

    sessions, current = [], None
    for r in range(3, ad.max_row + 1):
        name = ad.cell(r, 2).value
        if name:
            current = str(name).strip()
        if not isinstance(ad.cell(r, 3).value, int):
            continue                                    # 'Overall' or blank row

        delivered_puffs = ad.cell(r, 4).value           # D
        per_puff_ref = ad.cell(r, 5).value              # E
        delivered_time = ad.cell(r, 6).value            # F
        counted_puffs = ad.cell(r, 10).value            # J
        recorded_time = ad.cell(r, 11).value            # K

        if delivered_time is None and None not in (delivered_puffs, per_puff_ref):
            delivered_time = delivered_puffs * per_puff_ref
        if None in (delivered_puffs, delivered_time, counted_puffs, recorded_time):
            continue

        sessions.append(dict(device=current, session=ad.cell(r, 3).value,
                             delivered_puffs=delivered_puffs,
                             delivered_time=delivered_time,
                             counted_puffs=counted_puffs,
                             recorded_time=recorded_time))
    return included, sessions


def aggregate(sessions, included, level, all_devices):
    """Collapse sessions to the requested unit of analysis.

    Returns (labels, reference, friends) where reference and friends are the two
    methods' measurements in seconds.
    """
    keep = {n.lower() for n in included}
    if not all_devices:
        sessions = [s for s in sessions if s["device"].lower() in keep]

    if level == "session":
        groups = [(f"{s['device']} #{s['session']}", [s]) for s in sessions]
    else:
        order, buckets = [], {}
        for s in sessions:
            key = s["device"].lower()
            if key not in buckets:
                buckets[key] = []
                order.append((s["device"], key))
            buckets[key].append(s)
        groups = [(name, buckets[key]) for name, key in order]

    labels, ref, fri = [], [], []
    for name, rows in groups:
        d_puffs = sum(r["delivered_puffs"] for r in rows)
        d_time = sum(r["delivered_time"] for r in rows)
        c_puffs = sum(r["counted_puffs"] for r in rows)
        r_time = sum(r["recorded_time"] for r in rows)
        if MODE == "total":
            labels.append(name); ref.append(d_time); fri.append(r_time)
        else:
            # Duration agreement is defined over detected puffs; undetected
            # puffs are a counting error, not a duration error, and are
            # reported separately as precision/recall.
            if c_puffs == 0:
                print(f"  skipping {name}: no puffs detected, duration undefined",
                      file=sys.stderr)
                continue
            labels.append(name)
            ref.append(d_time / d_puffs)
            fri.append(r_time / c_puffs)
    return labels, np.asarray(ref, float), np.asarray(fri, float)


def bland_altman_stats(diff):
    """Bias, limits of agreement, and 95% CIs for each (Bland & Altman 1999)."""
    n = diff.size
    bias, sd = diff.mean(), diff.std(ddof=1)
    t = stats.t.ppf(0.975, n - 1)
    bias_ci = t * sd / np.sqrt(n)
    # SE of a limit of agreement, accounting for uncertainty in both bias and SD
    loa_se = sd * np.sqrt(1.0 / n + Z ** 2 / (2.0 * (n - 1)))
    return dict(n=n, bias=bias, sd=sd, bias_ci=bias_ci,
                upper=bias + Z * sd, lower=bias - Z * sd, loa_ci=t * loa_se)


def plot_total(mean, diff, st, out):
    fig, ax = plt.subplots(figsize=(9.6, 6.6), dpi=150)

    ax.axhspan(st["bias"] - st["bias_ci"], st["bias"] + st["bias_ci"],
               color=C_BIAS_BAND, zorder=1)
    for loa in (st["upper"], st["lower"]):
        ax.axhspan(loa - st["loa_ci"], loa + st["loa_ci"], color=C_LOA_BAND, zorder=1)

    ax.axhline(st["bias"], color=C_BIAS, lw=1.8, zorder=4,
               label=f"Bias = {st['bias']:+.2f} s")
    ax.axhline(st["upper"], color=C_LOA, ls="--", lw=1.5, zorder=4,
               label=f"+1.96 SD = {st['upper']:+.2f} s")
    ax.axhline(st["lower"], color=C_LOA, ls="--", lw=1.5, zorder=4,
               label=f"\u22121.96 SD = {st['lower']:+.2f} s")

    fit = stats.linregress(mean, diff)
    xs = np.linspace(mean.min(), mean.max(), 100)
    ax.plot(xs, fit.intercept + fit.slope * xs, color=C_FIT, ls="-.", lw=2.0,
            zorder=5,
            label=f"Prop. bias fit (r={fit.rvalue:.2f}, p={fit.pvalue:.3f})")

    ax.axhline(0, color="0.4", lw=0.8, ls=":", zorder=2)
    ax.scatter(mean, diff, s=95, color=C_POINT, edgecolor=C_POINT_EDGE,
               linewidth=0.8, zorder=6)

    ax.set_title("Bland\u2013Altman Plot: Total Atomizer Activation Time\n"
                 "(Puffing Machine vs FRIENDS)", fontsize=14, pad=12)
    ax.set_xlabel("Mean of Puffing Machine and FRIENDS total puff time (s)", fontsize=12)
    ax.set_ylabel("Difference: Puffing Machine \u2212 FRIENDS (s)", fontsize=12)
    ax.grid(True, ls=":", color="0.8", lw=0.7)
    ax.set_axisbelow(True)
    ax.legend(loc="upper left", framealpha=0.95, fontsize=10, edgecolor="0.3")
    fig.tight_layout()
    fig.savefig(out, bbox_inches="tight")
    print(f"  proportional bias slope = {fit.slope:+.5f} s per s "
          f"({100 * fit.slope:+.2f}% of measured duration)")
    return fit


def plot_perpuff(labels, diff, st, out):
    """Reference variance is zero, so difference vs mean is perfectly coupled.
    Plot the difference distribution against device index instead."""
    fig, ax = plt.subplots(figsize=(9.6, 6.6), dpi=150)
    x = np.arange(1, diff.size + 1)

    ax.axhspan(st["bias"] - st["bias_ci"], st["bias"] + st["bias_ci"],
               color=C_BIAS_BAND, zorder=1)
    for loa in (st["upper"], st["lower"]):
        ax.axhspan(loa - st["loa_ci"], loa + st["loa_ci"], color=C_LOA_BAND, zorder=1)

    ax.axhline(st["bias"], color=C_BIAS, lw=1.8, zorder=4,
               label=f"Bias = {st['bias']:+.2f} s")
    ax.axhline(st["upper"], color=C_LOA, ls="--", lw=1.5, zorder=4,
               label=f"+1.96 SD = {st['upper']:+.2f} s")
    ax.axhline(st["lower"], color=C_LOA, ls="--", lw=1.5, zorder=4,
               label=f"\u22121.96 SD = {st['lower']:+.2f} s")
    ax.axhline(0, color="0.4", lw=0.8, ls=":", zorder=2)
    ax.scatter(x, diff, s=95, color=C_POINT, edgecolor=C_POINT_EDGE,
               linewidth=0.8, zorder=6)

    ax.set_title("Bland\u2013Altman Analysis: Mean Puff Duration\n"
                 "(Puffing Machine vs FRIENDS)", fontsize=14, pad=12)
    ax.set_xlabel("ENDS device", fontsize=12)
    ax.set_ylabel("Difference: Puffing Machine \u2212 FRIENDS (s)", fontsize=12)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=60, ha="right", fontsize=8)
    ax.grid(True, ls=":", color="0.8", lw=0.7)
    ax.set_axisbelow(True)
    ax.legend(loc="upper left", framealpha=0.95, fontsize=10, edgecolor="0.3")
    ax.text(0.99, 0.02,
            "Machine reference fixed at 3.000 s (zero variance):\n"
            "difference vs mean is coupled by construction,\n"
            "so proportional bias is not estimable.",
            transform=ax.transAxes, ha="right", va="bottom", fontsize=8.5,
            color="0.35", style="italic")
    fig.tight_layout()
    fig.savefig(out, bbox_inches="tight")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--mode", choices=["total", "perpuff"], default="total")
    ap.add_argument("--level", choices=["device", "session"], default="device")
    ap.add_argument("--all-devices", action="store_true",
                    help="include the devices that lost detectability")
    ap.add_argument("-o", "--out", default=None)
    args = ap.parse_args()

    global MODE
    MODE = args.mode

    included, sessions = read_sessions(args.workbook)
    labels, ref, fri = aggregate(sessions, included, args.level, args.all_devices)

    diff = ref - fri
    mean = (ref + fri) / 2.0
    st = bland_altman_stats(diff)

    out = args.out or f"bland_altman_bench_{args.mode}_{args.level}.png"
    print(f"\nmode={args.mode}  level={args.level}  n={st['n']}  "
          f"devices retained={len(included)}")
    print(f"  bias  = {st['bias']:+.4f} s  (95% CI "
          f"{st['bias'] - st['bias_ci']:+.4f} to {st['bias'] + st['bias_ci']:+.4f})")
    print(f"  SD    = {st['sd']:.4f} s")
    print(f"  LoA   = {st['lower']:+.4f} to {st['upper']:+.4f} s "
          f"(each \u00b1{st['loa_ci']:.4f})")
    if args.mode == "total":
        plot_total(mean, diff, st, out)
    else:
        plot_perpuff(labels, diff, st, out)
    print(f"  wrote {out}")


if __name__ == "__main__":
    main()
